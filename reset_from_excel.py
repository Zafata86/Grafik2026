"""
Парсира ГРАФИК 2026.xlsx → презаписва init_data.py → нулира базата данни.
"""
import openpyxl
import sqlite3
import os
import sys

BASE = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE, 'database.db')
EXCEL = os.path.join(BASE, 'ГРАФИК 2026.xlsx')

YEAR = 2026

# Индекс на лист → месец (0-базирано)
# 0=януари1, 1=февруари(компактен), 2=март(компактен),
# 3=февруари(официален-skip), 4=март(официален-skip),
# 5=април, 6=май, 7=юни, 8=юли, 9=август,
# 10=септември, 11=октомври, 12=ноември, 13=декември
SHEET_IDX_TO_MONTH = {
    0: 1,   # януари
    1: 2,   # февруари (компактен)
    2: 3,   # март (компактен)
    # 3, 4 → дублирани, пропускаме
    5: 4,   # април
    6: 5,   # май
    7: 6,   # юни
    8: 7,   # юли
    9: 8,   # август
    10: 9,  # септември
    11: 10, # октомври
    12: 11, # ноември
    13: 12, # декември
}


def parse_code(val):
    """Конвертира стойност от Excel клетка към код за графика."""
    if val is None:
        return ''
    if isinstance(val, (int, float)):
        iv = int(val)
        if iv in (1, 2, 8, 0):
            return str(iv)
        return ''
    if isinstance(val, str):
        v = val.strip()
        if v in ('1', '2', '8', '0', 'П', 'Н', 'Б'):
            return v
        return ''
    return ''


def parse_sheet(ws):
    """
    Намира хедър реда с 'Р №' и извлича:
    [(tab_str, name, smyana, {day: code, ...}), ...]
    """
    header_row_idx = None
    header_vals = None
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if row[0] == 'Р №':
            header_row_idx = ridx
            header_vals = row
            break

    if header_row_idx is None:
        return []

    # Намери кои колони са дни (1..31)
    day_cols = {}  # col_index → day_number
    for i, h in enumerate(header_vals):
        if isinstance(h, (int, float)) and 1 <= int(h) <= 31:
            day_cols[i] = int(h)

    employees = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # Таб номер в колона 0
        tab_raw = row[0]
        if tab_raw is None:
            continue
        if isinstance(tab_raw, (int, float)):
            tab_str = str(int(tab_raw))
        else:
            tab_str = str(tab_raw).strip()
        if not tab_str.isdigit():
            continue

        name = str(row[1]).strip() if row[1] else ''
        smyana = str(row[2]).strip().upper() if row[2] else ''
        if not name or not smyana or smyana not in ('А','Б','В','Г','Р','Д','Е'):
            continue

        days = {}
        for col_i, day_num in day_cols.items():
            if col_i < len(row):
                code = parse_code(row[col_i])
                if code:
                    days[day_num] = code

        employees.append((tab_str, name, smyana, days))

    return employees


# ── 1. Парсирай Excel ─────────────────────────────────────────────────────────
print('Четене на ГРАФИК 2026.xlsx ...')
wb = openpyxl.load_workbook(EXCEL, data_only=True)
sheets = wb.sheetnames

all_entries = []   # [(tab, year, month, day, code, smyana, name)]
emp_info    = {}   # tab → (name, smyana)

for idx, month in SHEET_IDX_TO_MONTH.items():
    if idx >= len(sheets):
        print(f'  Предупреждение: лист {idx} не съществува!')
        continue
    ws = wb[sheets[idx]]
    emps = parse_sheet(ws)
    print(f'  Mesec {month:02d}: {sheets[idx]!r} -> {len(emps)} sluzhiteli')
    for tab, name, smyana, days in emps:
        emp_info[tab] = (name, smyana)
        for day, code in days.items():
            all_entries.append((tab, YEAR, month, day, code, smyana, name))

print(f'\nОбщо записи: {len(all_entries)}')
print(f'Служители:   {len(emp_info)}')
months_found = sorted(set(e[2] for e in all_entries))
print(f'Месеци:      {months_found}')

# ── 2. Провери дали служителите в Excel съвпадат с тези в DB ──────────────────
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
db_users = {str(r['tab_number']): r for r in conn.execute('SELECT * FROM users').fetchall()}

print('\n── Сравнение служители Excel vs БД ──')
missing_in_db  = []
extra_in_db    = []

for tab, (name, sm) in sorted(emp_info.items()):
    if tab not in db_users:
        missing_in_db.append(f'  ЛИПСВА В БД: {tab} {name} ({sm})')
    else:
        db_sm = db_users[tab]['smyana'].upper()
        if db_sm != sm:
            print(f'  СМЯНА различна: {tab} {name} → Excel={sm}, БД={db_sm}')

for tab, u in db_users.items():
    if tab not in emp_info:
        extra_in_db.append(f'  ИМА В БД, НЕ В EXCEL: {tab} {u["name"]}')

for m in missing_in_db:
    print(m)
for m in extra_in_db:
    print(m)

if missing_in_db:
    print('\nВнимание: има служители в Excel, които ги няма в БД!')
    print('Продължавам само с тези, които са и в двете.')

# ── 3. Нулиране на БД ─────────────────────────────────────────────────────────
print('\n── Нулиране на базата данни ──')

# Изтрий всички заявки
conn.execute('DELETE FROM vacation_requests')
print('  vacation_requests → изчистени')

conn.execute('DELETE FROM schedule_change_requests')
print('  schedule_change_requests → изчистени')

# Изтрий всички записи в графика
conn.execute('DELETE FROM schedule_entries')
print('  schedule_entries → изчистени')

# Нулирай vacation_days_remaining = vacation_days_total
conn.execute('UPDATE users SET vacation_days_remaining = vacation_days_total')
print('  vacation_days_remaining → нулирани до vacation_days_total')

# Изчисти lock/amendment флагове
conn.execute("DELETE FROM app_settings WHERE key LIKE 'sched_locked_%'")
conn.execute("DELETE FROM app_settings WHERE key LIKE 'amendment_needed_%'")
print('  lock/amendment флагове → изчистени')

# Изчисти миграционни флагове (за да може init_data да зареди наново)
conn.execute("DELETE FROM app_settings WHERE key IN ('reload_may_june_2026','plan_code_migration_done')")
print('  миграционни флагове → изчистени')

# ── 4. Зареди данните от Excel ────────────────────────────────────────────────
print('\n── Зареждане на нови данни ──')
inserted = 0
skipped  = 0

for tab, yr, mo, day, code, sm, name in all_entries:
    if tab not in db_users:
        skipped += 1
        continue

    emp_id = db_users[tab]['id']
    leave_status = 'approved' if code == '0' else 'normal'
    plan_code = code if code in ('1','2','8','0','Н','П') else None

    conn.execute(
        '''INSERT INTO schedule_entries
           (employee_id, year, month, day, code, leave_status, plan_code)
           VALUES (?, ?, ?, ?, ?, ?, ?)''',
        (emp_id, yr, mo, day, code, leave_status, plan_code)
    )
    inserted += 1

conn.commit()
conn.close()

print(f'  Вмъкнати:  {inserted}')
print(f'  Пропуснати (липсват в БД): {skipped}')

# ── 5. Изчисти временния файл ─────────────────────────────────────────────────
for f in ('_sheets.txt', '_sheets2.txt'):
    p = os.path.join(BASE, f)
    if os.path.exists(p):
        os.remove(p)

print('\n✓ Готово! Рестартирайте сървъра.')
