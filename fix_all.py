"""Нулира отпуски, изчиства заявки, сравнява и поправя графика спрямо Excel."""
import openpyxl, sqlite3, os

BASE  = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(BASE, 'ГРАФИК 2026.xlsx')
DB    = os.path.join(BASE, 'database.db')
YEAR  = 2026

SHEET_IDX_TO_MONTH = {0:1,1:2,2:3,5:4,6:5,7:6,8:7,9:8,10:9,11:10,12:11,13:12}

def parse_code(val):
    if val is None: return ''
    if isinstance(val,(int,float)):
        iv=int(val); return str(iv) if iv in(1,2,8,0) else ''
    if isinstance(val,str):
        v=val.strip(); return v if v in('1','2','8','0','П','Н','Б') else ''
    return ''

def parse_sheet(ws):
    hdr_idx, hdr = None, None
    for ridx,row in enumerate(ws.iter_rows(min_row=1,max_row=20,values_only=True),1):
        if row[0]=='Р №': hdr_idx,hdr=ridx,row; break
    if not hdr_idx: return []
    day_cols={i:int(h) for i,h in enumerate(hdr)
              if isinstance(h,(int,float)) and 1<=int(h)<=31}
    result=[]
    for row in ws.iter_rows(min_row=hdr_idx+1,values_only=True):
        tab=row[0]
        if tab is None: continue
        tab=str(int(tab)) if isinstance(tab,(int,float)) else str(tab).strip()
        if not tab.isdigit(): continue
        name  =str(row[1]).strip() if row[1] else ''
        smyana=str(row[2]).strip().upper() if row[2] else ''
        if not name or smyana not in('А','Б','В','Г','Р','Д','Е'): continue
        days={day_cols[i]:parse_code(row[i]) for i in day_cols
              if i<len(row) and parse_code(row[i])}
        result.append((tab,name,smyana,days))
    return result

# ── 1. Четем Excel ────────────────────────────────────────────────────────────
print('Четене на Excel...')
wb=openpyxl.load_workbook(EXCEL,data_only=True); sheets=wb.sheetnames

excel_data={}  # (tab,month,day) -> code
emp_info={}

for idx,month in SHEET_IDX_TO_MONTH.items():
    ws=wb[sheets[idx]]
    for tab,name,smyana,days in parse_sheet(ws):
        emp_info[tab]=(name,smyana)
        for day,code in days.items():
            excel_data[(tab,month,day)]=code

print(f'Excel: {len(excel_data)} записа, {len(emp_info)} служители')

# ── 2. Четем DB ───────────────────────────────────────────────────────────────
conn=sqlite3.connect(DB); conn.row_factory=sqlite3.Row

users   ={str(r['tab_number']):r for r in conn.execute('SELECT * FROM users')}
entries =conn.execute(
    'SELECT se.*,u.tab_number FROM schedule_entries se '
    'JOIN users u ON se.employee_id=u.id WHERE se.year=?',(YEAR,)
).fetchall()

db_data={}  # (tab,month,day) -> (entry_id,code,leave_status)
for e in entries:
    key=(str(e['tab_number']),e['month'],e['day'])
    db_data[key]=(e['id'],e['code'],e['leave_status'])

print(f'DB:    {len(db_data)} записа')

# ── 3. Сравнение ──────────────────────────────────────────────────────────────
print('\nСравнение Excel vs DB...')

only_excel   = []  # трябва да се добавят в DB
only_db      = []  # трябва да се изтрият от DB (не са в Excel)
diff_code    = []  # различен код — трябва да се поправи

all_keys = set(excel_data) | {k for k in db_data}

for key in sorted(all_keys):
    tab,mo,day=key
    if tab not in emp_info: continue  # непознат служител
    in_excel = key in excel_data
    in_db    = key in db_data

    if in_excel and not in_db:
        only_excel.append(key)
    elif in_db and not in_excel:
        db_code=db_data[key][1]
        # Ако е '0' (отпуска) — вероятно тестова заявка, трябва да се изтрие
        only_db.append(key)
    elif in_excel and in_db:
        ec=excel_data[key]
        dc=db_data[key][1]
        if ec!=dc:
            diff_code.append((key,ec,dc))

print(f'  Само в Excel (липсват в DB):    {len(only_excel)}')
print(f'  Само в DB (не са в Excel):      {len(only_db)}')
print(f'  Различен код:                   {len(diff_code)}')

if only_db:
    print('\n  Записи само в DB (ще се изтрият):')
    for tab,mo,day in only_db[:30]:
        code=db_data[(tab,mo,day)][1]
        name=emp_info.get(tab,(tab,''))[0] if tab in emp_info else users.get(tab,{}).get('name','?')
        print(f'    {tab} {name[:20]:<20} M{mo:02d}/D{day:02d} код={code!r}')
    if len(only_db)>30: print(f'    ... и още {len(only_db)-30}')

if diff_code:
    print('\n  Различни кодове (ще се поправят):')
    for (tab,mo,day),ec,dc in diff_code[:20]:
        name=emp_info.get(tab,(tab,''))[0]
        print(f'    {tab} {name[:20]:<20} M{mo:02d}/D{day:02d} DB={dc!r} -> Excel={ec!r}')

# ── 4. Прилагане на поправки ──────────────────────────────────────────────────
print('\nПрилагане...')

# 4a. Добави липсващи от Excel
for tab,mo,day in only_excel:
    eid=users[tab]['id']
    code=excel_data[(tab,mo,day)]
    ls='approved' if code=='0' else 'normal'
    pc=code if code in('1','2','8','0','Н','П') else None
    conn.execute(
        'INSERT OR REPLACE INTO schedule_entries'
        '(employee_id,year,month,day,code,leave_status,plan_code) VALUES(?,?,?,?,?,?,?)',
        (eid,YEAR,mo,day,code,ls,pc))

# 4b. Изтрий записи, които не са в Excel (включително тестови отпуски)
for tab,mo,day in only_db:
    eid=users[tab]['id']
    conn.execute(
        'DELETE FROM schedule_entries WHERE employee_id=? AND year=? AND month=? AND day=?',
        (eid,YEAR,mo,day))

# 4c. Поправи различните кодове
for (tab,mo,day),ec,dc in diff_code:
    eid=users[tab]['id']
    ls='approved' if ec=='0' else 'normal'
    pc=ec if ec in('1','2','8','0','Н','П') else None
    conn.execute(
        'UPDATE schedule_entries SET code=?,leave_status=?,plan_code=? '
        'WHERE employee_id=? AND year=? AND month=? AND day=?',
        (ec,ls,pc,eid,YEAR,mo,day))

print(f'  Добавени:  {len(only_excel)}')
print(f'  Изтрити:   {len(only_db)}')
print(f'  Поправени: {len(diff_code)}')

# ── 5. Нулиране заявки и отпуски ──────────────────────────────────────────────
conn.execute('DELETE FROM vacation_requests')
conn.execute('DELETE FROM schedule_change_requests')
conn.execute('UPDATE users SET vacation_days_remaining=20, vacation_days_total=20')
print('\n  vacation_requests     -> изчистени')
print('  schedule_change_requests -> изчистени')
print('  vacation_days (всички) -> 20')

# 6. Флагове
conn.execute("INSERT OR REPLACE INTO app_settings(key,value) VALUES('reload_may_june_2026','1')")
conn.execute("INSERT OR REPLACE INTO app_settings(key,value) VALUES('plan_code_migration_done','1')")
conn.execute("DELETE FROM app_settings WHERE key LIKE 'sched_locked_%'")
conn.execute("DELETE FROM app_settings WHERE key LIKE 'amendment_needed_%'")

conn.commit()

# ── 7. Финална статистика ─────────────────────────────────────────────────────
print('\nФинален брой записи по месеци:')
for m,c in conn.execute(
    'SELECT month,COUNT(*) FROM schedule_entries WHERE year=2026 GROUP BY month ORDER BY month'
).fetchall():
    print(f'  Месец {m:02d}: {c}')

total=conn.execute('SELECT COUNT(*) FROM schedule_entries WHERE year=2026').fetchone()[0]
print(f'  ОБЩО:    {total}')
conn.close()
print('\nГотово!')
