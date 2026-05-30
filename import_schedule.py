"""
Импорт на график Януари–Юни 2026 от ГРАФИК 2026.xlsx
"""
import openpyxl
import sqlite3
import os
import calendar

EXCEL = r'C:\Users\zafat\Documents\ГРАФИК 2026.xlsx'
DB    = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.db')

# Лист → (година, месец)
SHEETS = [
    ('януари1',     2026, 1),
    ('февруари 26', 2026, 2),
    ('март26  (2)', 2026, 3),
    ('април',       2026, 4),
    ('май',         2026, 5),
    ('юни',         2026, 6),
]

# Нормите са в ред 7 (0-indexed), колони 19 и 25
NORM_ROW = 7
NORM_COL_ABVG = 19
NORM_COL_RDE  = 25

# Приемливи кодове
VALID = {'1', '2', '8', '0', 'Б', 'Н', 'П'}

def normalize(v):
    if v is None:
        return ''
    s = str(v).strip().upper()
    if s in VALID:
        return s
    if s == 'НП':
        return 'Н'   # Неплатен → Неявяване
    return ''

def read_sheet(ws, year, month):
    """Връща list от dicts {tab, codes:{day:code}} + norm hours."""
    days_in_month = calendar.monthrange(year, month)[1]

    # Норма часове от ред 7
    norm_row_data = list(ws.iter_rows(min_row=NORM_ROW, max_row=NORM_ROW, values_only=True))[0]
    try:
        hours_abvg = int(norm_row_data[NORM_COL_ABVG])
    except (TypeError, ValueError, IndexError):
        hours_abvg = 160
    try:
        hours_rde = int(norm_row_data[NORM_COL_RDE])
    except (TypeError, ValueError, IndexError):
        hours_rde = hours_abvg

    # Служителски редове – всеки ред с цяло число в колона A
    employees = []
    seen_tabs = set()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=3 + days_in_month,
                            values_only=True):
        tab_val = str(row[0]).strip() if row[0] else ''
        try:
            int(tab_val)
        except ValueError:
            continue
        if tab_val in seen_tabs:
            continue
        seen_tabs.add(tab_val)

        codes = {}
        for day in range(1, days_in_month + 1):
            col_idx = 2 + day   # col A=0, B=1, C=2, D=3 → day1 at index 3
            if col_idx < len(row):
                code = normalize(row[col_idx])
                if code:
                    codes[day] = code

        employees.append({'tab': tab_val, 'codes': codes})

    return employees, hours_abvg, hours_rde


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Кеш tab_number → employee_id
    tab_map = {
        r['tab_number']: r['id']
        for r in cur.execute('SELECT id, tab_number FROM users').fetchall()
    }

    total_inserted = 0
    total_deleted  = 0

    for sheet_name, year, month in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f'  !! Лист не е намерен: {sheet_name}')
            continue

        ws = wb[sheet_name]
        employees, hours_abvg, hours_rde = read_sheet(ws, year, month)

        # Изтрий стари записи за месеца
        cur.execute(
            'DELETE FROM schedule_entries WHERE year=? AND month=?',
            (year, month)
        )
        deleted = cur.rowcount
        total_deleted += deleted

        # Задай норма часове
        cur.execute(
            'INSERT OR REPLACE INTO month_settings (year, month, hours_abvg, hours_rde) VALUES (?,?,?,?)',
            (year, month, hours_abvg, hours_rde)
        )

        inserted = 0
        missing  = 0
        for emp in employees:
            emp_id = tab_map.get(emp['tab'])
            if not emp_id:
                print(f'    !! Не намерен в БД: таб {emp["tab"]}')
                missing += 1
                continue
            for day, code in emp['codes'].items():
                leave_status = 'approved' if code == '0' else 'normal'
                cur.execute(
                    '''INSERT OR REPLACE INTO schedule_entries
                       (employee_id, year, month, day, code, leave_status)
                       VALUES (?, ?, ?, ?, ?, ?)''',
                    (emp_id, year, month, day, code, leave_status)
                )
                inserted += 1
        total_inserted += inserted

        month_bg = {1:'Януари',2:'Февруари',3:'Март',4:'Април',5:'Май',6:'Юни'}
        print(f'  {month_bg[month]:10} {year} | норма {hours_abvg}/{hours_rde} ч. | '
              f'изтрити {deleted:3} | внесени {inserted:3} записа | '
              f'{len(employees)} служители{" (!"+str(missing)+" липсват)" if missing else ""}')

    conn.commit()
    conn.close()
    print(f'\nОбщо внесени: {total_inserted} записа, изтрити стари: {total_deleted}')
    print('Готово!')


if __name__ == '__main__':
    main()
