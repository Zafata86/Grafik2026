"""Генерира init_data.py от ГРАФИК 2026.xlsx и презарежда DB."""
import openpyxl, sqlite3, os

BASE  = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(BASE, 'ГРАФИК 2026.xlsx')
YEAR  = 2026

SHEET_IDX_TO_MONTH = {0:1, 1:2, 2:3, 5:4, 6:5, 7:6,
                      8:7, 9:8, 10:9, 11:10, 12:11, 13:12}


def parse_code(val):
    if val is None: return ''
    if isinstance(val, (int, float)):
        iv = int(val)
        return str(iv) if iv in (1, 2, 8, 0) else ''
    if isinstance(val, str):
        v = val.strip()
        return v if v in ('1','2','8','0','П','Н','Б') else ''
    return ''


def parse_sheet(ws):
    header_row_idx = None
    header_vals    = None
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if row[0] == 'Р №':
            header_row_idx = ridx
            header_vals    = row
            break
    if header_row_idx is None:
        return []
    day_cols = {}
    for i, h in enumerate(header_vals):
        if isinstance(h, (int, float)) and 1 <= int(h) <= 31:
            day_cols[i] = int(h)
    employees = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        tab_raw = row[0]
        if tab_raw is None:
            continue
        tab_str = str(int(tab_raw)) if isinstance(tab_raw, (int, float)) else str(tab_raw).strip()
        if not tab_str.isdigit():
            continue
        name   = str(row[1]).strip() if row[1] else ''
        smyana = str(row[2]).strip().upper() if row[2] else ''
        if not name or smyana not in ('А','Б','В','Г','Р','Д','Е'):
            continue
        days = {}
        for col_i, day_num in day_cols.items():
            if col_i < len(row):
                code = parse_code(row[col_i])
                if code:
                    days[day_num] = code
        employees.append((tab_str, name, smyana, days))
    return employees


# ── Парсиране ─────────────────────────────────────────────────────────────────
print('Четене на Excel...')
wb      = openpyxl.load_workbook(EXCEL, data_only=True)
sheets  = wb.sheetnames

all_entries = []  # (tab, yr, mo, day, code, smyana, name)
emp_info    = {}  # tab -> (name, smyana)

for idx, month in SHEET_IDX_TO_MONTH.items():
    ws = wb[sheets[idx]]
    for tab, name, smyana, days in parse_sheet(ws):
        emp_info[tab] = (name, smyana)
        for day, code in days.items():
            all_entries.append((tab, YEAR, month, day, code, smyana, name))

print(f'Намерени: {len(all_entries)} записа, {len(emp_info)} служители')

# ── Генерирай init_data.py ────────────────────────────────────────────────────
print('Пишем init_data.py ...')
INIT_PATH = os.path.join(BASE, 'init_data.py')

header = '''\
# Генериран автоматично от ГРАФИК 2026.xlsx
import sqlite3, os

SCHEDULE_DATA = [
    # (tab_number, year, month, day, code, leave_status)
'''

footer = '''\
]


def load():
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.db')
    conn = sqlite3.connect(db_path)
    tab_map = {str(r[0]): r[1]
               for r in conn.execute('SELECT tab_number, id FROM users').fetchall()}
    for tab, yr, mo, day, code, ls in SCHEDULE_DATA:
        eid = tab_map.get(str(tab))
        if not eid:
            continue
        pc = code if code in ('1', '2', '8', '0', 'Н', 'П') else None
        conn.execute(
            """INSERT OR REPLACE INTO schedule_entries
               (employee_id, year, month, day, code, leave_status, plan_code)
               VALUES (?,?,?,?,?,?,?)""",
            (eid, yr, mo, day, code, ls, pc))
    conn.commit()
    conn.close()
    print(f'Заредени {len(SCHEDULE_DATA)} записа от init_data.')
'''

with open(INIT_PATH, 'w', encoding='utf-8') as f:
    f.write(header)
    for tab, yr, mo, day, code, sm, name in sorted(all_entries, key=lambda x: (x[0], x[2], x[3])):
        ls = 'approved' if code == '0' else 'normal'
        f.write(f"    ({tab!r}, {yr}, {mo:2d}, {day:2d}, {code!r}, {ls!r}),\n")
    f.write(footer)

print('init_data.py записан.')

# ── Презареди DB ──────────────────────────────────────────────────────────────
print('Презареждане на базата данни...')
DB_PATH = os.path.join(BASE, 'database.db')
conn    = sqlite3.connect(DB_PATH)

tab_map = {str(r[0]): r[1]
           for r in conn.execute('SELECT tab_number, id FROM users').fetchall()}

# Изчисти всички записи в графика
conn.execute('DELETE FROM schedule_entries')

inserted = 0
skipped  = 0
for tab, yr, mo, day, code, sm, name in all_entries:
    eid = tab_map.get(tab)
    if not eid:
        skipped += 1
        continue
    ls = 'approved' if code == '0' else 'normal'
    pc = code if code in ('1','2','8','0','Н','П') else None
    conn.execute(
        """INSERT INTO schedule_entries
           (employee_id, year, month, day, code, leave_status, plan_code)
           VALUES (?,?,?,?,?,?,?)""",
        (eid, yr, mo, day, code, ls, pc))
    inserted += 1

# Нулирай заявки и отпуски
conn.execute('DELETE FROM vacation_requests')
conn.execute('DELETE FROM schedule_change_requests')
conn.execute('UPDATE users SET vacation_days_remaining = vacation_days_total')
conn.execute("DELETE FROM app_settings WHERE key LIKE 'sched_locked_%'")
conn.execute("DELETE FROM app_settings WHERE key LIKE 'amendment_needed_%'")

# Постави флаговете за да не се презаписва при рестарт
conn.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES ('reload_may_june_2026','1')")
conn.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES ('plan_code_migration_done','1')")

conn.commit()
conn.close()

print(f'Вмъкнати: {inserted}, пропуснати: {skipped}')
print()

# Отпечатай по месеци
conn2 = sqlite3.connect(DB_PATH)
rows  = conn2.execute(
    'SELECT month, COUNT(*) FROM schedule_entries WHERE year=2026 GROUP BY month ORDER BY month'
).fetchall()
conn2.close()
for m, c in rows:
    print(f'  Месец {m:02d}: {c} записа')

print()
print('Готово! Рестартирайте сървъра.')
